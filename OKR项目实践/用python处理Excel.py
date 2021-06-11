#从命令参数中获取关键字,从sys.argv中读取命令行参数
#从requests模块取得查询结果页面
#找到每个查询结果的链接
#调用webbrowser.open()函数打开web浏览器

#web文件下载整理


















#用python处理Excel表
#! python3
# 匹配两个版本的崩溃信息，确认崩溃较早的引入版本.

import openpyxl
import os
import requests
from bs4 import BeautifulSoup

os.getcwd()
os.chdir('D:\workfile\study\python')
wb = openpyxl.workbook()
print("input the checking versions like '5192808' ")
version_a = input()
version_b = input()
print("ok,loading......")
wb.create_sheet(index=0,title=version_a)
wb.create_sheet(index=1,title=version_b)
wb.create_sheet(index=2,title='两个版本相同的崩溃信息')

#获取单个版本崩溃排名前10的崩溃信息数组
def getTop10Details(version):#参数是需要获取崩溃信息的版本号，类似5192109
    #动态设置url网页链接
    url = 'http://172.20.205.59/index.php?selectVersion=' + version + '&Submit=%E6%90%9C%E7%B4%A2'
    #定义空数组，用来存储后续输入的崩溃信息数组
    results=[]
    res = requests.get(url)
    res.raise_for_status()
    bf = BeautifulSoup(res.text,'lxml')
    #按照tag<td class='single-line single-line-td'>条件来查找所有符合的崩溃信息
    textx=bf.find_all('td', class_= 'single-line single-line-td')
    # 循环输出崩溃排名前10的崩溃信息
    for i in range(0,9):
        results.append(textx[i].text)
    return(results)

sheet_a = wb.get_sheet_by_name(version_a)
result_a = getTop10Details(version_a)
for i in range(1,len(result_a)+1):
    sheet_a.cell(row=i,column=5).value = result_a[i-1]
sheet_b = wb.get_sheet_by_name(version_b)

result_b = getTop10Details(version_b)
for i in range(1,len(result_a)+1):
    sheet_b.cell(row=i,column=5).value = result_a[i-1]




#比较两个数组相同的元素，输出仅有相同元素的新数组
def Compare(resultx,resulty):
    resultz=[]
    for i in range(0,len(resultx)):
        for j in range(0,len(resulty)):
            if resultx[i]==resulty[j]:
                resultz.append(resultx[i])
    return resultz

#分行标号打印数组
def TypeArray(array):
    for i in range(0,len(array)):
        print(str(i+1)+' ' +array[i])

#输入函数需要查询对比的两个版本号参数信息

result_c = Compare(getTop10Details(version_a),getTop10Details(version_b))
sheet_c = wb.get_sheet_by_name('两个版本相同的崩溃信息')
for i in range(1,len(result_a)+1):
    sheet_c.cell(row=i,column=5).value = result_a[i-1]
wb.save('版本对比.xlsx')
